#!/usr/bin/env python3
"""
DigiMinds Weekly Google Ads + GA4 Audit — 4 Clients
Ortho | HeatWeave | Tack Media | Doug PMF
Period: Last 7 days (pulled live from Windsor.ai)
Output: ~/Downloads/[Client]/[YYYY-MM-DD]/[HH-MM-SS]/ → XLSX + PDF + HTML
"""
import os, sys, json, datetime, math
from pathlib import Path
from collections import defaultdict

try:
    import requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ─── CONFIG ───────────────────────────────────────────────────────────────────
WINDSOR_API_KEY = os.environ.get("WINDSOR_API_KEY", "")
WINDSOR_BASE    = "https://connectors.windsor.ai/all"
DOWNLOADS       = Path.home() / "Downloads"
CACHE_DIR       = Path.home() / ".claude" / "cache" / "windsor"

TODAY     = datetime.date.today()
DATE_FROM = (TODAY - datetime.timedelta(days=7)).isoformat()
DATE_TO   = TODAY.isoformat()
PERIOD    = f"{DATE_FROM} → {DATE_TO} (Last 7 Days)"

# Windsor field sets — split because Google Ads can't mix impression + conversion metrics
GADS_PERF_FIELDS = ["date","account_id","account_name","campaign","campaign_status",
                    "campaign_bidding_strategy_type","spend","clicks","impressions",
                    "cpc","ctr","search_impression_share"]

GADS_CONV_FIELDS = ["date","account_id","campaign","conversions","all_conversions",
                    "phone_calls","cost_per_conversion","roas"]

GA4_FIELDS = ["date","account_id","sessions","active_users","conversions",
              "bounce_rate","average_session_duration","engaged_sessions","default_channel_group"]

# ─── 4 CLIENTS ────────────────────────────────────────────────────────────────
CLIENTS = {
    "Ortho": {
        "full_name":     "City Orthopedics & Sports Medicine",
        "color":         "#1a4f8a",
        "accent":        "#e8f0fb",
        "gads_accounts": ["106-254-8978"],
        "ga4_accounts":  ["258012138", "439336281"],
        "benchmarks": {
            "cpa_target":  80,
            "ctr_target":  0.05,
            "cpc_max":     15,
            "conv_rate_target": 0.08,
            "is_target":   0.40,
        },
        "compliance":  ["HIPAA", "FTC", "NJ-§13:35-6.10"],
        "status_note": "⚙️ Account in setup — 14-day pilot (kickoff May 12). No live campaigns yet.",
        "phone":       "201-500-9450 / 201-613-3388",
        "gtm":         "GTM-K96NFSF",
    },
    "HeatWeave": {
        "full_name":     "HeatWeave HVAC",
        "color":         "#c0392b",
        "accent":        "#fdf2f1",
        "gads_accounts": ["494-668-8111"],
        "ga4_accounts":  ["375400573"],
        "benchmarks": {
            "cpa_target":  150,
            "ctr_target":  0.04,
            "cpc_max":     35,
            "conv_rate_target": 0.05,
            "is_target":   0.25,
        },
        "compliance":  [],
        "status_note": "",
        "phone":       "",
        "gtm":         "",
    },
    "Tack": {
        "full_name":     "Tack Media",
        "color":         "#1e8449",
        "accent":        "#eafaf1",
        "gads_accounts": ["185-925-4010"],
        "ga4_accounts":  ["394685444"],
        "benchmarks": {
            "cpa_target":  80,
            "ctr_target":  0.05,
            "cpc_max":     12,
            "conv_rate_target": 0.05,
            "is_target":   0.25,
        },
        "compliance":  [],
        "status_note": "",
        "phone":       "",
        "gtm":         "",
    },
    "DougPMF": {
        "full_name":     "Doug PMF (Private Money Funding)",
        "color":         "#5b2c8d",
        "accent":        "#f5eef8",
        "gads_accounts": ["347-328-6476"],
        "ga4_accounts":  ["530251886"],
        "benchmarks": {
            "cpa_target":  600,
            "ctr_target":  0.04,
            "cpc_max":     10,
            "conv_rate_target": 0.02,
            "is_target":   0.25,
        },
        "compliance":  [],
        "status_note": "",
        "phone":       "",
        "gtm":         "",
    },
}

# ─── DATA HELPERS ─────────────────────────────────────────────────────────────
def safe(val, default=0.0):
    try:
        v = float(val) if val not in (None, "", "N/A") else default
        return v if not math.isnan(v) else default
    except:
        return default

def load_cache(fname):
    p = CACHE_DIR / fname
    if p.exists():
        with open(p) as f:
            data = json.load(f)
        rows = data.get("result", data) if isinstance(data, dict) else data
        return rows
    return []

def windsor_api(connector, fields, accounts, date_from, date_to):
    if not REQUESTS_OK or not WINDSOR_API_KEY:
        return []
    params = {
        "api_key":    WINDSOR_API_KEY,
        "connector":  connector,
        "date_from":  date_from,
        "date_to":    date_to,
        "fields":     ",".join(fields),
        "account_id": ",".join(accounts),
    }
    try:
        r = requests.get(WINDSOR_BASE, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        return data.get("data", data) if isinstance(data, dict) else data
    except Exception as e:
        print(f"  Windsor API error: {e}")
        return []

def merge_gads(perf_rows, conv_rows):
    """Merge performance + conversion datasets by (account_id, campaign, date)."""
    conv_map = {}
    for r in conv_rows:
        k = (str(r.get("account_id","")), r.get("campaign",""), r.get("date",""))
        conv_map[k] = r
    merged = []
    for r in perf_rows:
        k = (str(r.get("account_id","")), r.get("campaign",""), r.get("date",""))
        m = dict(r)
        c = conv_map.get(k, {})
        m["conversions"]         = safe(c.get("conversions", 0))
        m["all_conversions"]     = safe(c.get("all_conversions", 0))
        m["phone_calls"]         = safe(c.get("phone_calls", 0))
        m["cost_per_conversion"] = c.get("cost_per_conversion")
        m["roas"]                = safe(c.get("roas", 0))
        merged.append(m)
    return merged

def get_gads_data(all_accounts):
    """Pull or load Google Ads data — merged perf + conv."""
    if WINDSOR_API_KEY:
        print("  Pulling GAds perf…")
        perf = windsor_api("google_ads", GADS_PERF_FIELDS, all_accounts, DATE_FROM, DATE_TO)
        print("  Pulling GAds conv…")
        conv = windsor_api("google_ads", GADS_CONV_FIELDS, all_accounts, DATE_FROM, DATE_TO)
        return merge_gads(perf, conv)
    else:
        rows = load_cache("gads_weekly.json")
        print(f"  Cache: gads_weekly.json → {len(rows)} rows")
        return rows

def get_ga4_data(all_accounts):
    if WINDSOR_API_KEY:
        print("  Pulling GA4…")
        return windsor_api("googleanalytics4", GA4_FIELDS, all_accounts, DATE_FROM, DATE_TO)
    else:
        rows = load_cache("ga4_weekly.json")
        print(f"  Cache: ga4_weekly.json → {len(rows)} rows")
        return rows

# ─── ANALYSIS ─────────────────────────────────────────────────────────────────
def status_badge_html(status):
    s = (status or "").upper()
    if s == "ENABLED":
        return '<span style="background:#27ae60;color:white;padding:1px 7px;border-radius:3px;font-size:10px;font-weight:700">ENABLED</span>'
    elif s == "PAUSED":
        return '<span style="background:#e67e22;color:white;padding:1px 7px;border-radius:3px;font-size:10px;font-weight:700">PAUSED</span>'
    elif s == "REMOVED":
        return '<span style="background:#c0392b;color:white;padding:1px 7px;border-radius:3px;font-size:10px;font-weight:700">REMOVED</span>'
    return f'<span style="background:#95a5a6;color:white;padding:1px 7px;border-radius:3px;font-size:10px">{s or "UNKNOWN"}</span>'

def analyze_gads(rows, client_key, cfg):
    bench = cfg["benchmarks"]
    if not rows:
        note = cfg.get("status_note", "")
        return {"error": note if note else "No Google Ads data in period.", "status_note": note}

    # ── Aggregate totals across ALL campaigns (show everything, no filter) ──
    total_spend      = sum(safe(r.get("spend")) for r in rows)
    total_clicks     = sum(safe(r.get("clicks")) for r in rows)
    total_impr       = sum(safe(r.get("impressions")) for r in rows)
    total_conv       = sum(safe(r.get("conversions")) for r in rows)
    total_all_conv   = sum(safe(r.get("all_conversions")) for r in rows)
    total_phone      = sum(safe(r.get("phone_calls")) for r in rows)
    total_form_conv  = max(0.0, total_conv - total_phone)

    avg_cpc  = total_spend / total_clicks  if total_clicks  else 0
    avg_ctr  = total_clicks / total_impr   if total_impr    else 0
    avg_cpa  = total_spend / total_conv    if total_conv    else 0
    # roas: only use rows that have roas > 0
    roas_rows = [safe(r.get("roas")) for r in rows if safe(r.get("roas")) > 0]
    avg_roas = sum(roas_rows) / len(roas_rows) if roas_rows else 0

    # ── Per-campaign breakdown (all statuses shown) ──
    camp_map = defaultdict(lambda: {
        "spend":0,"clicks":0,"impressions":0,"conversions":0,
        "all_conversions":0,"phone_calls":0,
        "status":"","strategy":"","sync_note":"",
    })
    for r in rows:
        c = r.get("campaign","Unknown")
        camp_map[c]["spend"]         += safe(r.get("spend"))
        camp_map[c]["clicks"]        += safe(r.get("clicks"))
        camp_map[c]["impressions"]   += safe(r.get("impressions"))
        camp_map[c]["conversions"]   += safe(r.get("conversions"))
        camp_map[c]["all_conversions"]+= safe(r.get("all_conversions"))
        camp_map[c]["phone_calls"]   += safe(r.get("phone_calls"))
        if r.get("campaign_status"):
            camp_map[c]["status"]    = r["campaign_status"]
        if r.get("campaign_bidding_strategy_type"):
            camp_map[c]["strategy"]  = r["campaign_bidding_strategy_type"]
        if r.get("_note"):
            camp_map[c]["sync_note"] = r["_note"]

    campaigns = []
    for name, d in camp_map.items():
        d["name"]       = name
        d["form_conv"]  = max(0, d["conversions"] - d["phone_calls"])
        d["cpc"]        = d["spend"] / d["clicks"]       if d["clicks"]      else 0
        d["ctr"]        = d["clicks"] / d["impressions"] if d["impressions"] else 0
        d["cpa"]        = d["spend"] / d["conversions"]  if d["conversions"] else 0
        d["roas_calc"]  = 0  # roas at campaign level needs conv value — not available separately
        campaigns.append(d)
    campaigns.sort(key=lambda x: x["spend"], reverse=True)

    enabled_count = sum(1 for c in campaigns if c["status"].upper() == "ENABLED")
    paused_count  = sum(1 for c in campaigns if c["status"].upper() == "PAUSED")

    # ── Issue detection ──
    issues = []
    P = lambda p,i,d,f: {"priority":p,"issue":i,"detail":d,"fix":f}

    # Critical: zero conversions with real spend
    if total_conv == 0 and total_spend > 50:
        issues.append(P("CRITICAL","Zero tracked conversions",
            f"${total_spend:,.2f} spent — 0 conversions recorded (phone_calls={total_phone:.0f})",
            "Check conversion tag in GTM. Verify GA4 key events. Confirm call tracking is linked to Google Ads."))

    # Phone calls
    if total_phone == 0 and total_spend > 50:
        issues.append(P("HIGH","Zero phone call conversions",
            f"${total_spend:,.2f} spent — 0 calls tracked. Calls are primary intent signal for this client.",
            "Add call extension + call conversion in Google Ads. Link to GTM call tracking tag."))
    elif total_phone > 0:
        cpc_call = total_spend / total_phone
        if cpc_call > bench["cpa_target"]:
            issues.append(P("MEDIUM",f"Cost per call ${cpc_call:.0f} above target",
                f"{total_phone:.0f} calls at ${cpc_call:.0f}/call — target <${bench['cpa_target']}",
                "Tighten geo targeting, add negatives, improve ad relevance to reduce CPL."))

    # CPC
    if avg_cpc > bench["cpc_max"] and total_clicks > 10:
        issues.append(P("HIGH",f"CPC ${avg_cpc:.2f} above max ${bench['cpc_max']}",
            f"Avg CPC ${avg_cpc:.2f} across {total_clicks:.0f} clicks",
            "Add negatives, tighten match types, review bid strategy caps."))

    # CTR
    if total_impr > 300 and avg_ctr < 0.02:
        issues.append(P("HIGH","CTR critically low (<2%)",
            f"CTR {avg_ctr:.2%} on {total_impr:,.0f} impressions",
            "Rewrite headlines. Add callouts/sitelinks. Check keyword-ad relevance."))
    elif total_impr > 300 and avg_ctr < bench["ctr_target"]:
        issues.append(P("MEDIUM",f"CTR {avg_ctr:.2%} below target {bench['ctr_target']:.0%}",
            f"{total_impr:,.0f} impressions, {total_clicks:.0f} clicks",
            "Test RSA headline combinations. Pin best CTAs."))

    # CPA
    if avg_cpa > bench["cpa_target"] and total_conv > 0:
        issues.append(P("HIGH",f"CPA ${avg_cpa:.2f} above target ${bench['cpa_target']}",
            f"{total_conv:.1f} conversions at avg ${avg_cpa:.2f} each",
            "Pause low-conv campaigns. Improve LP conversion rate. Tighten targeting."))

    # Paused campaigns with spend (data anomaly / billing issue)
    paused_spend = sum(c["spend"] for c in campaigns if c["status"].upper() == "PAUSED")
    if paused_spend > 10:
        issues.append(P("MEDIUM",f"Paused campaigns have ${paused_spend:,.2f} spend",
            f"{paused_count} paused campaigns recorded spend this week — billing anomaly",
            "Verify in Google Ads UI — budget may have run before pause took effect. Check billing dates."))

    # Per-campaign flags
    for c in campaigns:
        if c["status"].upper() != "ENABLED":
            continue  # only flag enabled campaigns for campaign-level issues
        if c["impressions"] > 100 and c["ctr"] < 0.015:
            issues.append(P("HIGH",f"Low CTR: {c['name'][:40]}",
                f"CTR {c['ctr']:.2%} | {c['impressions']:,.0f} impr",
                "Rewrite ad copy, improve keyword match, add negatives."))
        if c["clicks"] > 20 and c["conversions"] == 0 and c["phone_calls"] == 0:
            issues.append(P("HIGH",f"No conv/calls: {c['name'][:40]}",
                f"${c['spend']:.0f} spent, {int(c['clicks'])} clicks — 0 form conv, 0 phone calls",
                "Check LP conversion tag. Review intent match. Pause if persists."))

    # Ortho-specific HIPAA
    if client_key == "Ortho" and total_spend > 0:
        issues.append(P("MEDIUM","HIPAA: Verify PHI exclusions active",
            "Enhanced Conversions active — confirm name/DOB/insurance excluded from data layer",
            f"Inspect GTM dataLayer in Tag Assistant. Review {cfg.get('gtm','')} config."))

    order = {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3}
    issues.sort(key=lambda x: order.get(x["priority"], 4))

    return {
        "total_spend":     total_spend,
        "total_clicks":    total_clicks,
        "total_impr":      total_impr,
        "total_conv":      total_conv,
        "total_all_conv":  total_all_conv,
        "total_phone":     total_phone,
        "total_form_conv": total_form_conv,
        "avg_cpc":         avg_cpc,
        "avg_ctr":         avg_ctr,
        "avg_cpa":         avg_cpa,
        "avg_roas":        avg_roas,
        "campaigns":       campaigns,
        "num_campaigns":   len(campaigns),
        "enabled_count":   enabled_count,
        "paused_count":    paused_count,
        "issues":          issues,
    }

def analyze_ga4(rows, client_key, cfg):
    if not rows:
        return {"error": "No GA4 data"}

    total_sessions = sum(safe(r.get("sessions")) for r in rows)
    total_users    = sum(safe(r.get("active_users")) for r in rows)
    total_conv     = sum(safe(r.get("conversions")) for r in rows)
    n              = len(rows)
    avg_bounce     = sum(safe(r.get("bounce_rate")) for r in rows) / n if n else 0
    avg_dur        = sum(safe(r.get("average_session_duration")) for r in rows) / n if n else 0

    channels = defaultdict(lambda: {"sessions":0,"conversions":0})
    for r in rows:
        ch = r.get("default_channel_group","Unknown")
        channels[ch]["sessions"]    += safe(r.get("sessions"))
        channels[ch]["conversions"] += safe(r.get("conversions"))

    paid_search_conv = channels.get("Paid Search",{}).get("conversions",0)
    paid_search_sess = channels.get("Paid Search",{}).get("sessions",0)

    issues = []
    P = lambda p,i,d,f: {"priority":p,"issue":i,"detail":d,"fix":f}

    if total_conv == 0 and total_sessions > 50:
        issues.append(P("CRITICAL","Zero GA4 conversions",
            f"{total_sessions:.0f} sessions — 0 conversions. Key events not configured.",
            "GA4 → Admin → Events → mark key_event: lead_form, phone_click, schedule_appointment."))

    if avg_bounce > 0.75:
        issues.append(P("HIGH","Bounce rate critically high",
            f"{avg_bounce:.1%} — very high drop-off. Likely LP/ad mismatch or slow page.",
            "Check page speed (Core Web Vitals). Match LP headline to ad copy. Add social proof above fold."))
    elif avg_bounce > 0.60:
        issues.append(P("MEDIUM","Elevated bounce rate",
            f"{avg_bounce:.1%} — benchmark <60%",
            "Improve content relevance. Add trust signals. Check mobile UX."))

    if avg_dur < 20:
        issues.append(P("HIGH","Session duration critically low (<20s)",
            f"Avg {avg_dur:.0f}s — bot traffic or instant bounces likely",
            "Check for spam/bot traffic in GA4. Verify tag fires correctly. Review LP load time."))
    elif avg_dur < 60:
        issues.append(P("MEDIUM","Low session duration",
            f"Avg {avg_dur:.0f}s — benchmark >60s",
            "Improve LP content depth. Add video. Fix mobile navigation."))

    if paid_search_sess > 20 and paid_search_conv == 0:
        issues.append(P("HIGH","Paid Search: 0 GA4 conversions",
            f"{paid_search_sess:.0f} paid search sessions — 0 GA4 conv. Tracking gap.",
            "Check GA4 conversion tag fires on thank-you/confirmation page after form submit."))

    order = {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3}
    issues.sort(key=lambda x: order.get(x["priority"], 4))

    return {
        "total_sessions": total_sessions,
        "total_users":    total_users,
        "total_conv":     total_conv,
        "avg_bounce":     avg_bounce,
        "avg_dur":        avg_dur,
        "channels":       dict(channels),
        "paid_conv":      paid_search_conv,
        "paid_sess":      paid_search_sess,
        "issues":         issues,
    }

# ─── XLSX ─────────────────────────────────────────────────────────────────────
def build_xlsx(client_key, cfg, gads, ga4, out_dir):
    wb  = openpyxl.Workbook()
    hex_color = cfg["color"].replace("#","")
    accent_hex = cfg["accent"].replace("#","")

    def hdr(ws, row, cols, title):
        ws.row_dimensions[row].height = 26
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", fgColor=hex_color)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).value = title

    # Summary
    ws = wb.active; ws.title = "Weekly Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    hdr(ws, 1, 2, f"{cfg['full_name']} — Weekly Audit ({DATE_FROM} → {DATE_TO})")

    row = 3
    def krow(label, val, fmt=None):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=fmt.format(val) if fmt and isinstance(val,(int,float)) else val)
        row += 1

    ws.cell(row=row, column=1, value="── GOOGLE ADS (Last 7 Days) ──").font = Font(bold=True, size=11, color=hex_color); row += 1
    if "error" in gads:
        krow("Status", gads.get("error","N/A"))
    else:
        krow("Total Spend",       gads["total_spend"],    "${:,.2f}")
        krow("Clicks",            gads["total_clicks"],   "{:,.0f}")
        krow("Impressions",       gads["total_impr"],     "{:,.0f}")
        krow("📞 Phone Calls",    gads["total_phone"],    "{:.0f}")
        krow("📋 Form Conv.",     gads["total_form_conv"],"{:.0f}")
        krow("Total Conv.",       gads["total_conv"],     "{:.1f}")
        krow("All Conv. (incl view-through)", gads["total_all_conv"], "{:.1f}")
        krow("Avg CPC",           gads["avg_cpc"],        "${:.2f}")
        krow("Avg CTR",           gads["avg_ctr"],        "{:.2%}")
        krow("Avg CPA",           gads["avg_cpa"],        "${:.2f}")
        krow("Campaigns (total)", gads["num_campaigns"],  "{:.0f}")
        krow("  → Enabled",       gads["enabled_count"],  "{:.0f}")
        krow("  → Paused",        gads["paused_count"],   "{:.0f}")

    row += 1
    ws.cell(row=row, column=1, value="── GA4 (Last 7 Days) ──").font = Font(bold=True, size=11, color=hex_color); row += 1
    if "error" in ga4:
        krow("Status", ga4["error"])
    else:
        krow("Sessions",          ga4["total_sessions"],  "{:,.0f}")
        krow("Active Users",      ga4["total_users"],     "{:,.0f}")
        krow("GA4 Conversions",   ga4["total_conv"],      "{:.0f}")
        krow("Paid Search Conv.", ga4["paid_conv"],       "{:.0f}")
        krow("Bounce Rate",       ga4["avg_bounce"],      "{:.1%}")
        krow("Avg Session Dur.",  ga4["avg_dur"],         "{:.0f}s")

    # Issues sheet
    ws2 = wb.create_sheet("Issues & Fixes")
    for col, w in zip("ABCDE", [12,38,44,56,10]):
        ws2.column_dimensions[col].width = w
    hdr(ws2, 1, 4, f"Issues & Prioritized Fixes — Week {DATE_FROM} → {DATE_TO}")
    for j, h in enumerate(["Priority","Issue","Detail","Fix"],1):
        c = ws2.cell(row=2, column=j, value=h)
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDDDDD")

    all_issues = gads.get("issues",[]) + ga4.get("issues",[])
    pc = {"CRITICAL":"FF4136","HIGH":"FF6B35","MEDIUM":"FFD700","LOW":"90EE90"}
    for i, iss in enumerate(all_issues, 3):
        ws2.cell(row=i, column=1, value=iss["priority"]).fill = PatternFill("solid", fgColor=pc.get(iss["priority"],"FFFFFF"))
        ws2.cell(row=i, column=1).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=iss["issue"])
        ws2.cell(row=i, column=3, value=iss["detail"])
        ws2.cell(row=i, column=4, value=iss["fix"])
        ws2.row_dimensions[i].height = 22
        for col in range(1,5):
            ws2.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="center")

    # Campaigns sheet — ALL campaigns, status labeled
    if "error" not in gads and gads.get("campaigns"):
        ws3 = wb.create_sheet("All Campaigns")
        hdrs3 = ["Campaign","Status","Spend","Clicks","Impr","📞 Calls","📋 Forms","Total Conv","All Conv","CPC","CTR","CPA"]
        col_ws3 = [42,10,12,10,10,10,10,10,10,10,10,12]
        for col, w in zip("ABCDEFGHIJKL", col_ws3):
            ws3.column_dimensions[col].width = w
        hdr(ws3, 1, len(hdrs3), f"All Campaigns — Last 7 Days (May 8–14, 2026) — All Statuses Shown")
        for j, h in enumerate(hdrs3,1):
            c = ws3.cell(row=2, column=j, value=h)
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DDDDDD")
        for i, camp in enumerate(gads["campaigns"], 3):
            call_c = camp.get("phone_calls", 0)
            form_c = camp.get("form_conv", 0)
            status = (camp["status"] or "").upper()
            row_data = [
                camp["name"], status,
                f"${camp['spend']:,.2f}", int(camp["clicks"]), int(camp["impressions"]),
                f"{call_c:.0f}", f"{form_c:.0f}", f"{camp['conversions']:.1f}",
                f"{camp['all_conversions']:.1f}",
                f"${camp['cpc']:.2f}" if camp["cpc"] else "—",
                f"{camp['ctr']:.2%}" if camp["ctr"] else "—",
                f"${camp['cpa']:.2f}" if camp["cpa"] else "—",
            ]
            for j, v in enumerate(row_data, 1):
                ws3.cell(row=i, column=j, value=v)
            # Color paused rows
            if status == "PAUSED":
                for j in range(1, len(hdrs3)+1):
                    ws3.cell(row=i, column=j).fill = PatternFill("solid", fgColor="FFF3CD")
            elif status == "ENABLED" and call_c == 0 and camp["spend"] > 20:
                ws3.cell(row=i, column=6).fill = PatternFill("solid", fgColor="FFE5E5")

    fname = out_dir / f"{client_key}_weekly_{TODAY}.xlsx"
    wb.save(fname); return fname

# ─── PDF ──────────────────────────────────────────────────────────────────────
def build_pdf(client_key, cfg, gads, ga4, out_dir):
    fname = out_dir / f"{client_key}_weekly_{TODAY}.pdf"
    PAGE_W, PAGE_H = A4
    LM = RM = 18*mm
    UW = PAGE_W - LM - RM

    doc = SimpleDocTemplate(str(fname), pagesize=A4,
                            leftMargin=LM, rightMargin=RM,
                            topMargin=18*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    brand  = colors.HexColor(cfg["color"])
    accent = colors.HexColor(cfg["accent"])

    H1   = ParagraphStyle("H1",   parent=styles["Heading1"], fontSize=16, textColor=brand, spaceAfter=2)
    H2   = ParagraphStyle("H2",   parent=styles["Heading2"], fontSize=11, textColor=brand, spaceAfter=2)
    BODY = ParagraphStyle("BODY", parent=styles["Normal"],   fontSize=9,  leading=13)
    SM   = ParagraphStyle("SM",   parent=styles["Normal"],   fontSize=7.5,leading=10, textColor=colors.grey)

    def ktbl(data, cw=None):
        cw = cw or [UW*0.42, UW*0.26, UW*0.32]
        t = Table(data, colWidths=cw, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), brand),
            ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, accent]),
            ("GRID",          (0,0),(-1,-1), 0.25, colors.lightgrey),
            ("TOPPADDING",    (0,0),(-1,-1), 3),
            ("BOTTOMPADDING", (0,0),(-1,-1), 3),
            ("LEFTPADDING",   (0,0),(-1,-1), 5),
        ]))
        return t

    story = []
    story.append(Paragraph("DigiMinds — Weekly PPC Audit", H1))
    story.append(Paragraph(cfg["full_name"], ParagraphStyle("SUB", parent=styles["Normal"], fontSize=12, textColor=colors.grey)))
    story.append(Paragraph(f"Period: {PERIOD} | Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", SM))
    if cfg.get("compliance"):
        story.append(Paragraph(f"Compliance: {' · '.join(cfg['compliance'])}", SM))
    story.append(HRFlowable(width=UW, color=brand, thickness=2, spaceAfter=5))

    # Google Ads KPIs
    story.append(Paragraph("Google Ads — Last 7 Days", H2))
    if "error" in gads:
        story.append(Paragraph(f"⚠️ {gads['error']}", BODY))
    else:
        bench = cfg["benchmarks"]
        def flag(val, target, low_is_bad=True):
            return ("✅" if val >= target else "⚠️") if low_is_bad else ("✅" if val <= target else "⚠️")

        kpi = [
            ["Metric", "Value", "Target / Status"],
            ["Total Spend",       f"${gads['total_spend']:,.2f}", "—"],
            ["Clicks",            f"{gads['total_clicks']:,.0f}", "—"],
            ["Impressions",       f"{gads['total_impr']:,.0f}",   "—"],
            ["📞 Phone Calls",    f"{gads['total_phone']:.0f}",
             f"{'✅ Tracking OK' if gads['total_phone']>0 else '⚠️ ZERO — check call tracking'}"],
            ["📋 Form Conv.",     f"{gads['total_form_conv']:.0f}", "—"],
            ["Total Conv.",       f"{gads['total_conv']:.1f}", "—"],
            ["All Conv. (view+click)", f"{gads['total_all_conv']:.1f}", "—"],
            ["Avg CPC",           f"${gads['avg_cpc']:.2f}",    f"{flag(gads['avg_cpc'],bench['cpc_max'],False)} ≤${bench['cpc_max']}"],
            ["Avg CTR",           f"{gads['avg_ctr']:.2%}",     f"{flag(gads['avg_ctr'],bench['ctr_target'])} {bench['ctr_target']:.0%}+"],
            ["Avg CPA",           f"${gads['avg_cpa']:.2f}" if gads['avg_cpa'] else "N/A",
             f"{flag(gads['avg_cpa'],bench['cpa_target'],False) if gads['avg_cpa'] else '—'} <${bench['cpa_target']}"],
            [f"Campaigns ({gads['enabled_count']} enabled / {gads['paused_count']} paused)",
             str(gads["num_campaigns"]), "—"],
        ]
        story.append(ktbl(kpi))
        story.append(Spacer(1, 5))

        # Campaign table
        if gads.get("campaigns"):
            story.append(Paragraph("Campaign Breakdown — All Statuses (Last 7 Days)", H2))
            cmp = [["Campaign","Status","Spend","Clicks","📞 Calls","📋 Forms","Total","CTR","CPA"]]
            for c in gads["campaigns"][:12]:
                flag_ctr = "🔴" if c["ctr"] < 0.015 else ("🟡" if c["ctr"] < bench["ctr_target"] else "🟢")
                cmp.append([
                    c["name"][:30], c["status"] or "?",
                    f"${c['spend']:,.0f}", f"{int(c['clicks'])}",
                    f"{c['phone_calls']:.0f}", f"{c['form_conv']:.0f}", f"{c['conversions']:.1f}",
                    f"{flag_ctr} {c['ctr']:.2%}" if c["ctr"] else "—",
                    f"${c['cpa']:.0f}" if c["cpa"] else "—",
                ])
            cw2 = [UW*0.25, UW*0.08, UW*0.09, UW*0.07, UW*0.08, UW*0.08, UW*0.07, UW*0.14, UW*0.14]
            t2 = Table(cmp, colWidths=cw2, repeatRows=1)
            ts2 = [
                ("BACKGROUND",    (0,0),(-1,0), brand),
                ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
                ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
                ("FONTSIZE",      (0,0),(-1,-1), 7),
                ("GRID",          (0,0),(-1,-1), 0.25, colors.lightgrey),
                ("TOPPADDING",    (0,0),(-1,-1), 2),
                ("BOTTOMPADDING", (0,0),(-1,-1), 2),
                ("LEFTPADDING",   (0,0),(-1,-1), 3),
            ]
            # Highlight paused rows
            for idx, c in enumerate(gads["campaigns"][:12], 1):
                if (c["status"] or "").upper() == "PAUSED":
                    ts2.append(("BACKGROUND", (0,idx), (-1,idx), colors.HexColor("#FFF3CD")))
            t2.setStyle(TableStyle(ts2))
            story.append(t2)

    story.append(Spacer(1, 6))

    # GA4
    story.append(Paragraph("GA4 Analytics — Last 7 Days", H2))
    if "error" in ga4:
        story.append(Paragraph(f"⚠️ {ga4['error']}", BODY))
    else:
        g4 = [
            ["Metric", "Value", "Benchmark"],
            ["Sessions",          f"{ga4['total_sessions']:,.0f}", "—"],
            ["Active Users",      f"{ga4['total_users']:,.0f}",   "—"],
            ["GA4 Conversions",   f"{ga4['total_conv']:.0f}",     "target >0"],
            ["Paid Search Conv.", f"{ga4['paid_conv']:.0f}",      "should match GAds conv."],
            ["Bounce Rate",       f"{ga4['avg_bounce']:.1%}",     "✅<60% / ⚠️>60%"],
            ["Avg Session Dur.",  f"{ga4['avg_dur']:.0f}s",       "target >60s"],
        ]
        story.append(ktbl(g4))

    story.append(Spacer(1, 6))

    # Issues
    story.append(HRFlowable(width=UW, color=brand, thickness=1, spaceAfter=4))
    all_issues = gads.get("issues",[]) + ga4.get("issues",[])
    story.append(Paragraph(f"Issues & Fixes — {len(all_issues)} found this week", H2))

    if not all_issues:
        story.append(Paragraph("✅ No critical issues detected.", BODY))
    else:
        pc = {"CRITICAL":colors.HexColor("#FF4136"),"HIGH":colors.HexColor("#FF6B35"),
              "MEDIUM":colors.HexColor("#FFD700"),"LOW":colors.HexColor("#90EE90")}
        iss_data  = [["#","Priority","Issue","Fix"]]
        row_fills = []
        for idx, iss in enumerate(all_issues, 1):
            iss_data.append([str(idx), iss["priority"], iss["issue"][:45], iss["fix"][:65]])
            row_fills.append(("BACKGROUND",(1,idx),(1,idx), pc.get(iss["priority"],colors.white)))
        cw3 = [UW*0.04, UW*0.11, UW*0.37, UW*0.48]
        t3  = Table(iss_data, colWidths=cw3, repeatRows=1)
        t3.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), brand),
            ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 7.5),
            ("GRID",          (0,0),(-1,-1), 0.25, colors.lightgrey),
            ("TOPPADDING",    (0,0),(-1,-1), 3),
            ("BOTTOMPADDING", (0,0),(-1,-1), 3),
            ("LEFTPADDING",   (0,0),(-1,-1), 4),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [colors.white, colors.HexColor("#f9f9f9")]),
        ] + row_fills))
        story.append(t3)

    story.append(Spacer(1,6))
    story.append(Paragraph(f"DigiMinds Weekly Audit | {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", SM))
    doc.build(story)
    return fname

# ─── HTML DASHBOARD ───────────────────────────────────────────────────────────
def build_html(client_key, cfg, gads, ga4, out_dir):
    color  = cfg["color"]
    accent = cfg["accent"]
    name   = cfg["full_name"]
    bench  = cfg["benchmarks"]
    all_issues = gads.get("issues",[]) + ga4.get("issues",[])

    priority_badge = {
        "CRITICAL": '<span style="background:#FF4136;color:white;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700">CRITICAL</span>',
        "HIGH":     '<span style="background:#FF6B35;color:white;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700">HIGH</span>',
        "MEDIUM":   '<span style="background:#FFD700;color:#333;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700">MEDIUM</span>',
        "LOW":      '<span style="background:#90EE90;color:#333;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:700">LOW</span>',
    }

    def kcard(label, val, sub="", warn=False, highlight=False):
        border = "border-left:4px solid #FF4136;" if warn else ("border-left:4px solid #27ae60;" if highlight else "")
        return f"""<div style="background:white;border-radius:8px;padding:14px 16px;box-shadow:0 2px 8px rgba(0,0,0,.08);min-width:120px;flex:1;{border}">
            <div style="font-size:10px;color:#888;text-transform:uppercase;letter-spacing:.5px">{label}</div>
            <div style="font-size:22px;font-weight:700;color:{color};margin:3px 0">{val}</div>
            <div style="font-size:10px;color:#aaa">{sub}</div></div>"""

    # ── Google Ads section ──
    if "error" in gads:
        gads_section = f"""<div style="background:#fff8e1;border:1px solid #FFD700;border-radius:8px;padding:18px;color:#666">
            ⚙️ {gads.get('error','')}{'<br><small>' + gads.get('status_note','') + '</small>' if gads.get('status_note') else ''}</div>"""
    else:
        ctr_warn  = gads["avg_ctr"]  < bench["ctr_target"]
        cpa_warn  = gads["avg_cpa"]  > bench["cpa_target"] and gads["avg_cpa"] > 0
        cpc_warn  = gads["avg_cpc"]  > bench["cpc_max"]
        call_warn = gads["total_phone"] == 0 and gads["total_spend"] > 50
        call_ok   = gads["total_phone"] > 0

        gads_cards = "".join([
            kcard("Spend",        f"${gads['total_spend']:,.0f}", f"{DATE_FROM} → {DATE_TO}"),
            kcard("Clicks",       f"{gads['total_clicks']:,.0f}", f"{gads['total_impr']:,.0f} impressions"),
            kcard("📞 Phone Calls", f"{gads['total_phone']:.0f}", "call conv tracked", call_warn, call_ok),
            kcard("📋 Form Conv.", f"{gads['total_form_conv']:.0f}", "form/web conv"),
            kcard("Total Conv.",  f"{gads['total_conv']:.1f}", "calls + forms"),
            kcard("Avg CPC",      f"${gads['avg_cpc']:.2f}",  f"target ≤${bench['cpc_max']}", cpc_warn),
            kcard("Avg CTR",      f"{gads['avg_ctr']:.2%}",   f"target {bench['ctr_target']:.0%}+", ctr_warn),
            kcard("Avg CPA",      f"${gads['avg_cpa']:.0f}" if gads['avg_cpa'] else "N/A",
                                  f"target <${bench['cpa_target']}", cpa_warn),
        ])

        # Campaign rows — ALL statuses, phone calls shown per campaign
        camp_rows = ""
        for c in gads["campaigns"]:
            call_c   = c.get("phone_calls", 0)
            form_c   = c.get("form_conv", 0)
            status   = (c.get("status") or "").upper()
            flag_ctr = "🔴" if c["ctr"] < 0.015 else ("🟡" if c["ctr"] < bench["ctr_target"] else "🟢") if c["ctr"] else "—"
            row_bg   = "#FFF3CD" if status == "PAUSED" else ("#fff3f3" if (c["conversions"]==0 and c["phone_calls"]==0 and c["clicks"]>15) else "white")
            call_style = 'style="color:#FF4136;font-weight:700"' if (call_c==0 and c["spend"]>20 and status=="ENABLED") else 'style="color:#27ae60;font-weight:700"' if call_c > 0 else ""

            sync_note = c.get("sync_note", "")
            name_cell = f'{c["name"][:50]}<br><small style="color:#e67e22;font-weight:400">{sync_note}</small>' if sync_note else c['name'][:50]
            camp_rows += f"""<tr style="background:{row_bg}">
                <td style="padding:8px 12px;font-size:13px;font-weight:600">{name_cell}</td>
                <td style="padding:8px 8px">{status_badge_html(status)}</td>
                <td style="padding:8px 12px">${c['spend']:,.2f}</td>
                <td style="padding:8px 12px">{int(c['clicks'])}</td>
                <td style="padding:8px 12px;font-size:13px" {call_style}>📞 {call_c:.0f}</td>
                <td style="padding:8px 12px">📋 {form_c:.0f}</td>
                <td style="padding:8px 12px"><strong>{c['conversions']:.1f}</strong></td>
                <td style="padding:8px 12px">{f"{flag_ctr} {c['ctr']:.2%}" if c['ctr'] else "—"}</td>
                <td style="padding:8px 12px">${f"{c['cpa']:.0f}" if c['cpa'] else "—"}</td>
            </tr>"""

        status_summary = f'<span style="font-size:11px;color:#555">✅ {gads["enabled_count"]} enabled &nbsp;|&nbsp; ⏸ {gads["paused_count"]} paused &nbsp;|&nbsp; All campaigns shown</span>'
        gads_section = f"""
        <div class="kpi-row">{gads_cards}</div>
        <div style="margin:12px 0 6px;display:flex;align-items:center;justify-content:space-between">
          <div class="section-title" style="margin:0;border:none">Campaign Breakdown — Last 7 Days</div>
          {status_summary}
        </div>
        <table>
          <thead><tr>
            <th>Campaign</th><th>Status</th><th>Spend</th><th>Clicks</th>
            <th>📞 Calls</th><th>📋 Forms</th><th>Total Conv.</th><th>CTR</th><th>CPA</th>
          </tr></thead>
          <tbody>{camp_rows}</tbody>
        </table>
        <div style="font-size:10px;color:#aaa;margin-top:4px">🟡 Paused rows in yellow. 📞 red = 0 calls with spend (tracking gap).</div>"""

    # ── GA4 section ──
    if "error" in ga4:
        ga4_section = f'<div style="color:#888;padding:16px">{ga4["error"]}</div>'
    else:
        bounce_warn = ga4["avg_bounce"] > 0.60
        dur_warn    = ga4["avg_dur"] < 60
        conv_warn   = ga4["total_conv"] == 0
        paid_warn   = ga4["paid_sess"] > 20 and ga4["paid_conv"] == 0
        ga4_cards   = "".join([
            kcard("Sessions",       f"{ga4['total_sessions']:,.0f}", "last 7 days"),
            kcard("Active Users",   f"{ga4['total_users']:,.0f}"),
            kcard("GA4 Conv.",      f"{ga4['total_conv']:.0f}", "all channels", conv_warn),
            kcard("Paid Search Conv.", f"{ga4['paid_conv']:.0f}", f"{ga4['paid_sess']:.0f} paid sessions", paid_warn),
            kcard("Bounce Rate",    f"{ga4['avg_bounce']:.1%}", "target <60%", bounce_warn),
            kcard("Avg Session",    f"{ga4['avg_dur']:.0f}s",  "target >60s", dur_warn),
        ])

        # Channel breakdown table
        ch_rows = ""
        for ch, d in sorted(ga4["channels"].items(), key=lambda x: x[1]["sessions"], reverse=True):
            cr = d["conversions"] / d["sessions"] if d["sessions"] else 0
            ch_rows += f"""<tr>
                <td style="padding:7px 12px">{ch}</td>
                <td style="padding:7px 12px">{d['sessions']:,.0f}</td>
                <td style="padding:7px 12px">{d['conversions']:.0f}</td>
                <td style="padding:7px 12px">{cr:.2%}</td>
            </tr>"""

        ga4_section = f"""<div class="kpi-row">{ga4_cards}</div>
        <div class="section-title" style="margin-top:12px">Channel Performance</div>
        <table>
          <thead><tr><th>Channel</th><th>Sessions</th><th>Conversions</th><th>Conv. Rate</th></tr></thead>
          <tbody>{ch_rows}</tbody>
        </table>"""

    # Issues table
    issues_html = ""
    for iss in all_issues:
        issues_html += f"""<tr>
            <td style="padding:10px 12px">{priority_badge.get(iss['priority'],'')}</td>
            <td style="padding:10px 12px;font-weight:600;font-size:13px">{iss['issue']}</td>
            <td style="padding:10px 12px;color:#555;font-size:12px">{iss['detail']}</td>
            <td style="padding:10px 12px;color:#1a4f8a;font-size:12px">💡 {iss['fix']}</td>
        </tr>"""
    if not issues_html:
        issues_html = '<tr><td colspan="4" style="padding:20px;color:green;text-align:center;font-weight:600">✅ No critical issues detected this week</td></tr>'

    compliance_badge = " ".join([
        f'<span style="background:{color};color:white;padding:2px 7px;border-radius:3px;font-size:10px">{c}</span>'
        for c in cfg.get("compliance",[])
    ])

    crit_count = sum(1 for i in all_issues if i["priority"]=="CRITICAL")
    high_count = sum(1 for i in all_issues if i["priority"]=="HIGH")
    alert_color = "#FF4136" if crit_count else ("#FF6B35" if high_count else "#27ae60")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{name} — Weekly Audit {TODAY}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;color:#222}}
  .header{{background:{color};color:white;padding:20px 28px}}
  .header h1{{font-size:19px;font-weight:700}}
  .header p{{font-size:11px;opacity:.8;margin-top:3px}}
  .section{{margin:16px 24px}}
  .section-title{{font-size:13px;font-weight:700;color:{color};margin-bottom:8px;padding-bottom:3px;border-bottom:2px solid {color}}}
  .kpi-row{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:10px}}
  table{{width:100%;border-collapse:collapse;background:white;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.06)}}
  thead tr{{background:{color};color:white}}
  th{{padding:9px 12px;text-align:left;font-size:11px;font-weight:600;letter-spacing:.3px}}
  tbody tr:nth-child(even){{background:{accent}}}
  tbody tr:hover{{background:#e8f4fd}}
  td{{font-size:12px}}
  .footer{{text-align:center;color:#bbb;font-size:11px;padding:18px}}
  .alert-bar{{background:white;margin:12px 24px;padding:12px 18px;border-radius:8px;box-shadow:0 1px 4px rgba(0,0,0,.06);display:flex;gap:16px;align-items:center;flex-wrap:wrap;border-left:4px solid {alert_color}}}
</style>
</head>
<body>
<div class="header">
  <h1>📊 {name} — Weekly PPC Audit</h1>
  <p>📅 {PERIOD} &nbsp;|&nbsp; Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} &nbsp;|&nbsp; {compliance_badge if compliance_badge else 'DigiMinds'}</p>
</div>

<div class="alert-bar">
  <span style="font-weight:700;font-size:13px">Issues this week: <span style="color:{alert_color}">{len(all_issues)}</span>
  {f"&nbsp;|&nbsp; 🔴 {crit_count} CRITICAL" if crit_count else ""}
  {f"&nbsp;|&nbsp; 🟠 {high_count} HIGH" if high_count else ""}
  </span>
  {"".join([f'<span style="font-size:11px">{priority_badge[i["priority"]]} {i["issue"][:45]}</span>' for i in all_issues[:3]])}
  {'<span style="font-size:11px;color:#888">+ more below ↓</span>' if len(all_issues) > 3 else ''}
</div>

<div class="section" style="margin-top:14px">
  <div class="section-title">🎯 Google Ads KPIs — Last 7 Days</div>
  {gads_section}
</div>

<div class="section">
  <div class="section-title">📈 GA4 Analytics — Last 7 Days</div>
  {ga4_section}
</div>

<div class="section">
  <div class="section-title">⚠️ All Issues & Prioritized Fixes ({len(all_issues)} total)</div>
  <table>
    <thead><tr><th style="width:10%">Priority</th><th style="width:24%">Issue</th><th style="width:30%">Detail</th><th style="width:36%">Fix</th></tr></thead>
    <tbody>{issues_html}</tbody>
  </table>
</div>

<div class="footer">DigiMinds Weekly Audit System | {TODAY} | Ortho · HeatWeave · Tack Media · Doug PMF</div>
</body></html>"""

    fname = out_dir / f"{client_key}_dashboard_{TODAY}.html"
    fname.write_text(html, encoding="utf-8")
    return fname

# ─── RUNNER ───────────────────────────────────────────────────────────────────
def run_client(client_key, cfg, all_gads, all_ga4):
    now      = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H-%M-%S")
    folder   = cfg["full_name"].replace(" ","_").replace("(","").replace(")","").replace(",","").replace("&","and")
    out_dir  = DOWNLOADS / folder / date_str / time_str
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*62}\n{cfg['full_name']}")
    print(f"Output: {out_dir}")

    gads_rows = [r for r in all_gads if str(r.get("account_id","")).replace("-","") in
                 [a.replace("-","") for a in cfg["gads_accounts"]]]
    ga4_rows  = [r for r in all_ga4  if str(r.get("account_id","")) in cfg["ga4_accounts"]]
    print(f"  GAds rows: {len(gads_rows)} | GA4 rows: {len(ga4_rows)}")

    gads = analyze_gads(gads_rows, client_key, cfg)
    ga4  = analyze_ga4(ga4_rows,  client_key, cfg)

    n_issues = len(gads.get("issues",[])) + len(ga4.get("issues",[]))
    print(f"  Issues: {n_issues}")

    xlsx = build_xlsx(client_key, cfg, gads, ga4, out_dir)
    pdf  = build_pdf( client_key, cfg, gads, ga4, out_dir)
    html = build_html(client_key, cfg, gads, ga4, out_dir)
    print(f"  ✓ {xlsx.name} | {pdf.name} | {html.name}")
    return {"xlsx":xlsx, "pdf":pdf, "html":html, "issues":n_issues, "dir":out_dir}

def main():
    print(f"DigiMinds Weekly Ads Audit — {TODAY}")
    print(f"Period: {PERIOD}")
    mode = "API" if WINDSOR_API_KEY else "CACHE (gads_weekly.json / ga4_weekly.json)"
    print(f"Data source: {mode}")

    all_gads_ids = [a for c in CLIENTS.values() for a in c["gads_accounts"]]
    all_ga4_ids  = [a for c in CLIENTS.values() for a in c["ga4_accounts"]]

    if WINDSOR_API_KEY:
        all_gads = get_gads_data(all_gads_ids)
        all_ga4  = get_ga4_data(all_ga4_ids)
    else:
        all_gads = get_gads_data(all_gads_ids)
        all_ga4  = get_ga4_data(all_ga4_ids)

    results = {}
    for key, cfg in CLIENTS.items():
        try:
            results[key] = run_client(key, cfg, all_gads, all_ga4)
        except Exception as e:
            import traceback; traceback.print_exc()
            results[key] = {"error": str(e)}

    print(f"\n{'='*62}\nWEEKLY AUDIT COMPLETE — {TODAY}")
    for key, r in results.items():
        if "error" in r:
            print(f"  ✗ {CLIENTS[key]['full_name']}: {r['error']}")
        else:
            print(f"  ✓ {CLIENTS[key]['full_name']}: {r['issues']} issues → {r['dir']}")

if __name__ == "__main__":
    main()
